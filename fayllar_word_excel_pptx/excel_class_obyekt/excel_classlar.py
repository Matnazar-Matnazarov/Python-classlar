from openpyxl import *
class Excel:
    def __init__(self, name):
        self.name=name
        self.wb=Workbook()
        self.ws=self.wb.active
    def read_text(self): print("\n".join(map(str,[ ' '.join(map(str,[i.value for i in j])) for j in self.ws.iter_rows()])))
    def write_cell(self, address, s):self.ws[address] = s
    def add(self, *s):  self.ws.append(list(s))
    def qushish(self, a, b, c):
        self.ws[c] = self.ws[a].value + self.ws[b].value
    def search(self, matn): return ' '.join(map(str,[''.join(map(str,[j.coordinate for j in i if j.value==matn])) for i in self.ws.iter_rows()]))
    def massiv(self, a, b):return [[j.value for j in i]  for i in self.ws[a:b]]
    def del_cell(self, a):    self.ws[a] = None
    def __repr__(self):return f"Fayl nomi: {self.name}"
    def __call__(self, a):return self.ws[a].value
    def saqlash(self, s):self.wb.save(s)

# quyidagi saytlardan o'rganib yozdim
# https://pythonbasics.org/read-excel/
# https://www.geeksforgeeks.org/reading-excel-file-using-python/
# https://www.dataquest.io/blog/reading-excel-file-python/
# https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/