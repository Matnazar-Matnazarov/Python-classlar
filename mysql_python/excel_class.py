from openpyxl import *
class Excel:
    def __init__(self, name=None,list_nomi=None):
        if name:
            self.wb=load_workbook(name)
            self.ws=self.wb[list_nomi]
        else:
            self.name = name
            self.wb = Workbook()
            self.ws = self.wb.active
    def read_text(self):
        return  ("\n".join(map(str,[ ' '.join(map(str,[i.value for i in j])) for j in self.ws.iter_rows()])))
    def add(self, d,table_nomi,ustun_nomlari):
        self.ws.append(ustun_nomlari)
        for i in d:
            self.ws.append(i)
        self.saqlash(f'{table_nomi}.xlsx')
    def matrix(self):
        return  [[i.value for i in j] for j in self.ws.iter_rows()]
    def write_cell(self, address, s):
        self.ws[address] = s
    def saqlash(self, s):
        self.wb.save(s)
